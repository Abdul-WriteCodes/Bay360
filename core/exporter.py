import io
import csv
import json
from typing import Any


DISPLAY_COLUMNS = [
    ("author_year", "Author & Year"),
    ("title", "Title"),
    ("research_context", "Research Context"),
    ("methodology", "Methodology"),
    ("independent_variables", "Independent Variables"),
    ("dependent_variable", "Dependent Variable"),
    ("control_variables", "Control Variables"),
    ("findings", "Key Findings"),
    ("theoretical_contributions", "Theoretical Contributions"),
    ("practical_contributions", "Practical Contributions"),
    ("strengths", "Strengths"),
    ("limitations", "Limitations"),
]


def papers_to_csv(papers: list[dict]) -> bytes:
    """Export papers to CSV bytes."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([col[1] for col in DISPLAY_COLUMNS])

    for p in papers:
        row = [p.get(col[0], "") for col in DISPLAY_COLUMNS]
        writer.writerow(row)

    return output.getvalue().encode("utf-8")


def papers_to_excel(papers: list[dict]) -> bytes:
    """Export papers to Excel bytes using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EmpiricX Results"

    # Style definitions
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a1e28", end_color="1a1e28", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
    )

    # Write headers
    for col_idx, (key, label) in enumerate(DISPLAY_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 36

    # Write data
    for row_idx, paper in enumerate(papers, 2):
        for col_idx, (key, _) in enumerate(DISPLAY_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=paper.get(key, ""))
            cell.alignment = cell_align
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=9)

        # Alternate row shading
        if row_idx % 2 == 0:
            for col_idx in range(1, len(DISPLAY_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="F8F9FF", end_color="F8F9FF", fill_type="solid"
                )

    # Column widths
    col_widths = [18, 30, 22, 18, 22, 18, 18, 35, 30, 30, 22, 25]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def papers_to_latex(papers: list[dict]) -> str:
    """Export papers to LaTeX longtable format."""

    def escape_latex(text: str) -> str:
        if not isinstance(text, str):
            text = str(text)
        replacements = [
            ("&", r"\&"),
            ("%", r"\%"),
            ("$", r"\$"),
            ("#", r"\#"),
            ("_", r"\_"),
            ("{", r"\{"),
            ("}", r"\}"),
            ("~", r"\textasciitilde{}"),
            ("^", r"\textasciicircum{}"),
            ("\\", r"\textbackslash{}"),
        ]
        for old, new in replacements:
            text = text.replace(old, new)
        return text

    lines = [
        r"\documentclass[12pt]{article}",
        r"\usepackage{longtable}",
        r"\usepackage{booktabs}",
        r"\usepackage{array}",
        r"\usepackage{geometry}",
        r"\geometry{margin=1in}",
        r"\usepackage{hyperref}",
        r"\renewcommand{\arraystretch}{1.4}",
        r"\begin{document}",
        r"\section*{Empirical Summary Table}",
        r"\begin{longtable}{p{2cm}p{3cm}p{2.5cm}p{2.5cm}p{4cm}p{3cm}}",
        r"\toprule",
        r"\textbf{Author \& Year} & \textbf{Research Context} & \textbf{Methodology} & "
        r"\textbf{Key Variables} & \textbf{Findings} & \textbf{Limitations} \\",
        r"\midrule",
        r"\endhead",
    ]

    for p in papers:
        ivs = escape_latex(p.get("independent_variables", ""))
        dv = escape_latex(p.get("dependent_variable", ""))
        variables = f"IV: {ivs}\\newline DV: {dv}"

        row_parts = [
            escape_latex(p.get("author_year", "")),
            escape_latex(p.get("research_context", "")),
            escape_latex(p.get("methodology", "")),
            variables,
            escape_latex(p.get("findings", "")),
            escape_latex(p.get("limitations", "")),
        ]
        lines.append(" & ".join(row_parts) + r" \\")
        lines.append(r"\midrule")

    lines += [
        r"\bottomrule",
        r"\end{longtable}",
        r"\end{document}",
    ]

    return "\n".join(lines)


def papers_to_json(papers: list[dict]) -> str:
    """Export papers to clean JSON string."""
    clean = []
    for p in papers:
        clean.append({k: v for k, v in p.items() if not k.startswith("_")})
    return json.dumps(clean, indent=2, ensure_ascii=False)
